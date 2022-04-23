# THIS SCRIPT IS FOR INTERNAL USE ONLY.

# First you need to generate a report from Odoo.
# "Sales" > open some sale order > "Moves" > select all records > "Action" >  "Export"
# Then in the templates select "Shipment amount based on date and stage" and hit the "Export" bottom.

# Now you can run the script

import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Counting the existing reports (Windows)
working_path = os.environ["USERPROFILE"] + "\\Downloads\\"
count = 0
for file in os.listdir(working_path):
    if(file.startswith("stock.picking")):
        count += 1

# Creating a new directory for output files
new_path = working_path + "\\گزارش فروش\\" 
if not os.path.exists(new_path):
    os.makedirs(new_path)

# Setting the output file names accordingly
if count == 1:
    input_file = working_path + "stock.picking.xlsx"
    output_file = new_path + "گزارش فروش.xlsx"
else:
    input_file = working_path + "stock.picking (%i).xlsx" % (count-1)
    output_file = new_path + "گزارش فروش (%i).xlsx"  % (count-1)

# Loading the latest excel file
workbook = load_workbook(input_file)
sheet = workbook.active

# Saving the stage name
if "مرحله" in sheet["B2"].value:
    stage = sheet["B2"].value
elif "مرحله" in sheet["I2"].value:
    stage = sheet["I2"].value
else:
    stage = ""
    sheet["B2"].value = None

sheet.delete_cols(9,10)
rows = sheet.max_row
for row in range(1, rows+1):
    if sheet["A%i" % row].value == None:
        sheet.delete_rows(row)
# Counting the rows
row_count = sheet.max_row

# Excel columns
columns = ["A","B","C","D","E","F","G","H"]

# Setting up the titles
sheet["A1"] = "شماره سفارش"
sheet["B1"] = "توضیحات"
sheet["B2"].value = stage
sheet["C1"] = "مشتری"
sheet["E1"] = "تاریخ ارسال"
sheet["G1"] = "مقدار کل سفارش"
sheet["H1"] = "مقدار ارسال شده"
sheet["G%i" % (row_count+2)] = ":اضافه ارسالی مربوط به مرحله ی قبل"
sheet["H%i" % (row_count+2)]=":جمع کل ارسالی"


# Clearing the duplicate values
for column in columns[:4]:
    for i in range(3, row_count+1):
        sheet["%s%i" % (column, i)] = ""

# Consuming each shipment amount from the total order
for i in range(3, row_count+1):
    sheet["G%i" % i].value = sheet["G%i" % (i-1)].value - sheet["H%i" % (i-1)].value
sheet["G%i" % (row_count+1)].value = sheet["G%i" % row_count].value - sheet["H%i" % row_count].value

# Checking if the shipment amount is equal to the order amount
if sheet["G%i" % (row_count+1)].value == 0:
    sheet["G%i" % (row_count+3)].value = 0
elif sheet["G%i" % (row_count+1)].value < 0:
    # If the shipment amount is more than order, we note it down as "Over Shipment"
    sheet["G%i" % (row_count+3)].value = abs(sheet["G%i" % (row_count+1)].value)
    sheet["G%i" % (row_count+1)].value = 0
else:
    sheet["G%i" % (row_count+3)].value = 0


# Sum of the total shipment
total_shipment = 0
for i in range(2, row_count+1):
    total_shipment += sheet["H%s" % i].value
sheet["H%i" % (row_count+3)].value = total_shipment

# Colorizing cells
for column in columns:
    sheet["%s1" % column].fill = PatternFill("solid", start_color="ddd9c4")
    sheet["%s%i" % (column, (row_count+2))].fill = PatternFill("solid", start_color="92cddc")
    sheet["%s%i" % (column, (row_count+3))].fill = PatternFill("solid", start_color="92cddc")
    for i in range(2, row_count+2):
        sheet["%s%i" % (column, (i))].fill = PatternFill("solid", start_color="ccffff")
        sheet["A%i" % i].fill = PatternFill("solid", start_color="da9694")

# Saving the output file
out_length = len(output_file)
if stage != "":
    workbook.save(output_file[:(out_length-5)] + " (%s)" % stage + ".xlsx")
else:
    workbook.save(output_file[:(out_length-5)] + ".xlsx")
print("Operation successful")
